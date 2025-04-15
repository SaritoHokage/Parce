from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment


options = webdriver.ChromeOptions()
options.add_argument('--ignore-certificate-errors')  
options.add_argument('--headless')
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")  # Установка User-Agent

driver = webdriver.Chrome(options=options)


base_url = "https://fedresurs.ru/biddings?tradeType=all&price=null&tradePeriod=null&bankrupt=null&onlyAvailableToParticipate=true&regionNumber=54&limit=15"


def get_lots_data():
    driver.get(base_url)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".u-card-result_mb")))  # Ждем загрузки лотов

    lot_data = []
    processed_lots = 0 
    page_number = 1 

    while True:
        
        lots = driver.find_elements(By.CSS_SELECTOR, ".u-card-result_mb")
        if not lots:
            print("Лоты не найдены. Завершаем обработку.")
            break
        
        for i in range(len(lots)):
            try:
                lot = lots[i]  

                
                WebDriverWait(lot, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".info-link-container a")))
                
                
                lot_link_element = lot.find_element(By.CSS_SELECTOR, ".info-link-container a")
                
                if not lot_link_element:
                    print("Не удалось получить элемент ссылки на лот. Пропускаем лот.")
                    continue
                
                print(f"Обработка лота № {processed_lots + 1}")  
                
                
                driver.execute_script("arguments[0].scrollIntoView();", lot_link_element)
                time.sleep(1)  

                
                driver.execute_script("arguments[0].click();", lot_link_element)
                time.sleep(5)  

                
                driver.switch_to.window(driver.window_handles[-1])

                
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.lot-item-tradeobject")))  
                
                
                lot_info = {
                    "Данные о лоте": driver.find_element(By.CSS_SELECTOR, "div.lot-item-tradeobject").text.strip(),
                    "Начальная цена": driver.find_element(By.CSS_SELECTOR, "body > fedresurs-app > app-standard-layout > div > bidding-item > div > bidding-card > div > div.information-content > bidding-card-lots > information-page-item > div > skeleton-loader > div > bidding-lot-card > div > div > div.lot-item-description > div:nth-child(1) > div.info-item-value").text.strip(),
                    "Прием заявок": driver.find_element(By.CSS_SELECTOR, "body > fedresurs-app > app-standard-layout > div > bidding-item > div > bidding-card > div > div.information-content > bidding-card-main > information-page-item > div > skeleton-loader > div > div:nth-child(3) > div.info-item-value").text.strip(),
                    "Вид торгов": driver.find_element(By.CSS_SELECTOR, "body > fedresurs-app > app-standard-layout > div > bidding-item > div > bidding-card > div > div.information-content > bidding-card-main > information-page-item > div > skeleton-loader > div > div:nth-child(2) > div.info-item-value").text.strip(),
                    "Ссылка на лот": driver.current_url  
                }
                lot_data.append(lot_info)

                
                driver.close()
                driver.switch_to.window(driver.window_handles[0]) 

                processed_lots += 1  

                
                if processed_lots % 15 == 0:
                    print(f"Обработано {processed_lots} лотов. Сохраняем данные в Excel на листе {page_number}.")
                    save_to_excel(lot_data, page_number)  
                    lot_data = []  
                    page_number += 1  

                
                if processed_lots % 15 == 0:
                    print("Обработано 15 лотов. Загружаем еще.")
                    try:
                        load_button = driver.find_element(By.CSS_SELECTOR, "body > fedresurs-app > app-standard-layout > div > bidding-search > div.u-form-item.u-form-item__wrapper_r-height > div > loader > div:nth-child(1) > bidding-search-result > el-tab-panel > div.tab-content > el-tab.selected > div > biddings-search-tab > loader > div.load-info > div > div")
                        driver.execute_script("arguments[0].scrollIntoView();", load_button)  
                        time.sleep(1) 
                        load_button.click()
                        time.sleep(5)  

                        
                        lots = driver.find_elements(By.CSS_SELECTOR, ".u-card-result_mb")
                        if not lots:
                            print("Новые лоты не найдены после загрузки. Завершаем обработку.")
                            return lot_data
                    except Exception as e:
                        print(f"Не удалось кликнуть на элемент загрузки. Ошибка: {e}")
                        return lot_data

            except Exception as e:
                print(f"Не удалось получить данные с лота. Ошибка: {e}")

    
    if lot_data:
        save_to_excel(lot_data, page_number)

    return lot_data


def save_to_excel(data, page_number):
    try:
        df = pd.DataFrame(data)
        filename = 'lots_data.xlsx'
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a' if os.path.exists(filename) else 'w') as writer:
            df.to_excel(writer, sheet_name=f'Page_{page_number}', index=False)

        
        wb = load_workbook(filename)
        ws = wb[f'Page_{page_number}']

        
        ws.column_dimensions['A'].width = 100  # Длина 1 столбца
        ws.column_dimensions['B'].width = 20   # Длина 2 столбца
        ws.column_dimensions['C'].width = 32   # Длина 3 столбца
        ws.column_dimensions['D'].width = 30   # Длина 4 столбца
        ws.column_dimensions['E'].width = 20   # Длина 5 столбца

        
        ws.row_dimensions[1].height = 15  # Высота первой строки

        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 100  # Фиксированная высота

        wb.save(filename)  
        wb.close()  
    except PermissionError:
        print("Ошибка: Не удалось сохранить файл 'lots_data.xlsx'. Убедитесь, что файл не открыт в Excel или у вас есть права на запись.")

# Основной цикл
def main():
    all_lots_data = []
    
    while True:
        lots_data = get_lots_data()
        all_lots_data.extend(lots_data)
        
        
        save_to_excel(all_lots_data, 1)  
        
        
        time.sleep(10)  

if __name__ == "__main__":
    main()
